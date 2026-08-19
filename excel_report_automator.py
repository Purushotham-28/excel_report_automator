#!/usr/bin/env python3
"""
Excel Report Automator
----------------------
A Python CLI tool that automates data ingestion, cleaning, validation,
summary metrics calculation, and formatted Excel report generation.

Author: Antigravity AI
"""

import sys
import os
import argparse
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_arguments():
    """Parse and validate command line arguments."""
    parser = argparse.ArgumentParser(
        description="Automated Data Cleaning, Validation, and Excel Report Generator",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument(
        "-i", "--input",
        required=True,
        help="Path to the input CSV or Excel file (.csv, .xlsx, .xls)"
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="Path for the generated output Excel report (.xlsx). Defaults to '[input_name]_report.xlsx'"
    )
    parser.add_argument(
        "--required-cols",
        nargs="+",
        default=None,
        help="Specific column names that must not be blank (space or comma separated)"
    )
    parser.add_argument(
        "--category-col",
        default=None,
        help="Column name to group by for summary calculations (auto-detected if not specified)"
    )
    return parser.parse_args()


def validate_input_filepath(file_path):
    """Validate existence, format, and size of input file."""
    if not os.path.exists(file_path):
        print(f"Error: Input file '{file_path}' does not exist.", file=sys.stderr)
        sys.exit(1)

    if not os.path.isfile(file_path):
        print(f"Error: Input path '{file_path}' is not a file.", file=sys.stderr)
        sys.exit(1)

    valid_extensions = (".csv", ".xlsx", ".xls")
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in valid_extensions:
        print(
            f"Error: Unsupported file format '{ext}'. Allowed formats: {', '.join(valid_extensions)}",
            file=sys.stderr
        )
        sys.exit(1)

    if os.path.getsize(file_path) == 0:
        print(f"Error: Input file '{file_path}' is empty (0 bytes).", file=sys.stderr)
        sys.exit(1)


def read_data_file(file_path):
    """Read CSV or Excel file into a pandas DataFrame with error handling."""
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == ".csv":
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
    except pd.errors.EmptyDataError:
        print(f"Error: CSV file '{file_path}' contains no data.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error: Failed to read file '{file_path}': {e}", file=sys.stderr)
        sys.exit(1)

    if df.empty:
        print(f"Error: File '{file_path}' loaded an empty dataset.", file=sys.stderr)
        sys.exit(1)

    return df


def standardize_column_name(name):
    """Clean and standardize a single column header name."""
    s = str(name).strip().lower()
    # Replace spaces and punctuation with underscores
    clean_chars = [c if c.isalnum() else "_" for c in s]
    s = "".join(clean_chars)
    # Collapse multiple consecutive underscores
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def standardize_columns(df):
    """Standardize column names of the DataFrame."""
    new_columns = [standardize_column_name(c) for c in df.columns]
    # Ensure unique column names if duplicates exist
    seen = {}
    final_columns = []
    for col in new_columns:
        if col in seen:
            seen[col] += 1
            final_columns.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            final_columns.append(col)
    df.columns = final_columns
    return df


def detect_category_column(df, user_category_col=None):
    """Auto-detect or validate the group-by category column."""
    if user_category_col:
        std_user_col = standardize_column_name(user_category_col)
        if std_user_col in df.columns:
            return std_user_col
        else:
            print(
                f"Warning: Specified category column '{user_category_col}' not found. Auto-detecting...",
                file=sys.stderr
            )

    # Keywords preference
    preferred_keywords = ["category", "region", "department", "type", "group", "product", "segment"]
    for kw in preferred_keywords:
        matching = [col for col in df.columns if kw in col]
        if matching:
            return matching[0]

    # Fallback to string/object column with distinct non-trivial values
    string_cols = [c for c in df.columns if df[c].dtype == "object" or df[c].dtype == "string" or df[c].dtype == "category"]
    for col in string_cols:
        nunique = df[col].nunique(dropna=True)
        if 1 < nunique <= len(df) // 2 or nunique <= 20:
            return col

    # Fallback to first non-numeric column or first column
    non_numeric = [col for col in df.columns if not pd.api.types.is_numeric_dtype(df[col])]
    if non_numeric:
        return non_numeric[0]

    return df.columns[0]


def clean_and_validate(df, required_cols_input=None):
    """
    Remove exact duplicate rows, standardize string fields, and validate rows.
    Returns:
        cleaned_df (DataFrame)
        errors_df (DataFrame)
        stats (dict)
    """
    total_raw_rows = len(df)

    # Clean strings: strip whitespace and treat whitespace-only strings as NaN
    str_columns = [col for col in df.columns if not pd.api.types.is_numeric_dtype(df[col]) and not pd.api.types.is_datetime64_any_dtype(df[col])]
    for col in str_columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace(r"^\s*$", np.nan, regex=True)
        # Fix string 'nan' / 'None' literals if converted from nulls
        df[col] = df[col].replace(["nan", "None", "NAN"], np.nan)

    # Deduplication
    df_dedup = df.drop_duplicates().copy()
    duplicate_rows_count = total_raw_rows - len(df_dedup)

    # Identify required columns
    if required_cols_input:
        required_cols = [standardize_column_name(c) for c in required_cols_input if standardize_column_name(c) in df.columns]
    else:
        # Default required columns: all non-numeric or key columns
        required_cols = list(df_dedup.columns)

    # Determine numeric columns for negative value checks
    numeric_cols = []
    for col in df_dedup.columns:
        # Try converting to numeric to check if column is numeric
        converted = pd.to_numeric(df_dedup[col], errors="coerce")
        if converted.notna().sum() > 0 and not pd.api.types.is_datetime64_any_dtype(df_dedup[col]):
            numeric_cols.append(col)
            # Standardize numeric column type
            df_dedup[col] = converted

    # Row-by-row validation
    error_rows = []
    cleaned_rows = []

    for idx, row in df_dedup.iterrows():
        reasons = []

        # Check 1: Missing required fields
        for col in required_cols:
            val = row[col]
            if pd.isna(val) or val is None or str(val).strip() == "":
                reasons.append(f"Missing required field '{col}'")

        # Check 2: Negative numbers in numeric columns
        for col in numeric_cols:
            val = row[col]
            if pd.notna(val):
                try:
                    num_val = float(val)
                    if num_val < 0:
                        reasons.append(f"Negative value in '{col}' ({num_val})")
                except (ValueError, TypeError):
                    reasons.append(f"Invalid non-numeric value in '{col}' ({val})")

        row_dict = row.to_dict()
        if reasons:
            row_dict["error_reason"] = "; ".join(reasons)
            error_rows.append(row_dict)
        else:
            cleaned_rows.append(row_dict)

    cleaned_df = pd.DataFrame(cleaned_rows, columns=df_dedup.columns) if cleaned_rows else pd.DataFrame(columns=df_dedup.columns)
    
    error_columns = list(df_dedup.columns) + ["error_reason"]
    errors_df = pd.DataFrame(error_rows, columns=error_columns) if error_rows else pd.DataFrame(columns=error_columns)

    stats = {
        "total_raw_rows": total_raw_rows,
        "duplicate_rows_count": duplicate_rows_count,
        "unique_rows_processed": len(df_dedup),
        "cleaned_rows_count": len(cleaned_df),
        "error_rows_count": len(errors_df),
        "numeric_cols": numeric_cols
    }

    return cleaned_df, errors_df, stats


def generate_summary_dataframes(cleaned_df, stats, category_col):
    """
    Generate summary stats: overall KPIs and grouped category stats.
    """
    # 1. Overall KPI Summary
    kpi_data = [
        {"Metric": "Total Raw Rows Input", "Value": stats["total_raw_rows"]},
        {"Metric": "Duplicate Rows Removed", "Value": stats["duplicate_rows_count"]},
        {"Metric": "Unique Rows Processed", "Value": stats["unique_rows_processed"]},
        {"Metric": "Valid Cleaned Rows", "Value": stats["cleaned_rows_count"]},
        {"Metric": "Flagged Error Rows", "Value": stats["error_rows_count"]},
        {
            "Metric": "Error Rate (%)",
            "Value": f"{(stats['error_rows_count'] / stats['unique_rows_processed'] * 100):.2f}%" if stats['unique_rows_processed'] > 0 else "0.00%"
        }
    ]
    kpi_df = pd.DataFrame(kpi_data)

    # 2. Category Grouped Breakdown
    if not cleaned_df.empty and category_col in cleaned_df.columns:
        num_cols = stats["numeric_cols"]
        
        # Build aggregations dict
        agg_dict = {cleaned_df.columns[0]: "count"}
        for num_c in num_cols:
            if num_c != category_col:
                agg_dict[num_c] = ["sum", "mean"]

        # Handle case where category_col might be the only non-numeric column
        grouped = cleaned_df.groupby(category_col, dropna=False)
        
        # Helper for clean header titles
        def get_stat_header(num_c, stat_type):
            clean_name = num_c.replace("_", " ").title()
            if stat_type == "sum":
                if clean_name.lower().startswith("total "):
                    return clean_name
                return f"Total {clean_name}"
            else:  # mean
                if clean_name.lower().startswith("total "):
                    clean_name = clean_name[6:]
                return f"Avg {clean_name}"

        # Build clean summary table
        summary_rows = []
        for cat, group in grouped:
            cat_label = "Unspecified" if pd.isna(cat) or str(cat).strip() == "" else str(cat)
            row = {"Category": cat_label, "Record Count": len(group)}
            
            for num_c in num_cols:
                if num_c != category_col:
                    sum_header = get_stat_header(num_c, "sum")
                    mean_header = get_stat_header(num_c, "mean")
                    row[sum_header] = group[num_c].sum()
                    row[mean_header] = group[num_c].mean()
            summary_rows.append(row)

        category_summary_df = pd.DataFrame(summary_rows)

        # Calculate Grand Totals row
        total_row = {"Category": "Grand Total", "Record Count": len(cleaned_df)}
        for num_c in num_cols:
            if num_c != category_col:
                sum_header = get_stat_header(num_c, "sum")
                mean_header = get_stat_header(num_c, "mean")
                total_row[sum_header] = cleaned_df[num_c].sum()
                total_row[mean_header] = cleaned_df[num_c].mean()
        
        category_summary_df = pd.concat([category_summary_df, pd.DataFrame([total_row])], ignore_index=True)
    else:
        category_summary_df = pd.DataFrame(columns=["Category", "Record Count", "Notes"])
        category_summary_df.loc[0] = ["N/A", 0, "No valid data available for category summary."]

    return kpi_df, category_summary_df


def create_formatted_excel(output_path, cleaned_df, errors_df, kpi_df, category_summary_df, category_col):
    """Write DataFrame sheets and format with openpyxl styling."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write basic dataframes first
        cleaned_df.to_excel(writer, sheet_name="Cleaned Data", index=False)
        errors_df.to_excel(writer, sheet_name="Errors", index=False)
        
        # Write Summary sheet manually for custom multi-table layout
        summary_ws = writer.book.create_sheet(title="Summary")

    # Re-open with openpyxl for detailed custom formatting
    wb = openpyxl.load_workbook(output_path)
    
    # Ensure sheet order: Cleaned Data, Errors, Summary
    # Remove default sheet if pandas created extra sheets
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Setup styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark Navy
    section_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")  # Slate Blue
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")    # Soft Blue
    error_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")    # Light Coral

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    regular_font = Font(name="Calibri", size=11)

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thick_bottom_side = Side(border_style="medium", color="1F4E79")
    double_bottom_side = Side(border_style="double", color="1F4E79")

    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    total_border = Border(top=thin_border_side, bottom=double_bottom_side, left=thin_border_side, right=thin_border_side)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # -------------------------------------------------------------
    # Format "Cleaned Data" & "Errors" Sheets
    # -------------------------------------------------------------
    for sheet_name in ["Cleaned Data", "Errors"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True

        if ws.max_row < 1 or ws.max_column < 1:
            continue

        # Header formatting
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill if sheet_name == "Cleaned Data" else PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = header_font
            cell.alignment = center_align
            cell.border = header_border

        # Data formatting
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = cell_border
                
                # Check data format
                val = cell.value
                col_name = str(ws.cell(row=1, column=col_idx).value or "").lower()
                
                if isinstance(val, (int, float)):
                    cell.alignment = right_align
                    if "price" in col_name or "amount" in col_name or "total" in col_name or "cost" in col_name or "sales" in col_name:
                        cell.number_format = "$#,##0.00"
                    elif isinstance(val, float):
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "#,##0"
                else:
                    cell.alignment = left_align

        ws.row_dimensions[1].height = 25
        auto_fit_columns(ws)

    # -------------------------------------------------------------
    # Build & Format "Summary" Sheet
    # -------------------------------------------------------------
    ws_sum = wb["Summary"]
    ws_sum.views.sheetView[0].showGridLines = True

    # Title Block
    ws_sum.cell(row=1, column=1, value="Data Automation & Analysis Report").font = title_font
    ws_sum.row_dimensions[1].height = 28

    # Section 1: KPI Summary Table
    ws_sum.cell(row=3, column=1, value="Processing Overview").font = Font(name="Calibri", size=12, bold=True, color="2F5597")
    
    kpi_headers = ["Metric", "Value"]
    for c_idx, h in enumerate(kpi_headers, 1):
        cell = ws_sum.cell(row=4, column=c_idx, value=h)
        cell.fill = section_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = header_border
    ws_sum.row_dimensions[4].height = 22

    curr_row = 5
    for _, kpi_row in kpi_df.iterrows():
        ws_sum.row_dimensions[curr_row].height = 20
        c1 = ws_sum.cell(row=curr_row, column=1, value=kpi_row["Metric"])
        c2 = ws_sum.cell(row=curr_row, column=2, value=kpi_row["Value"])
        
        c1.font = bold_font
        c1.border = cell_border
        c1.alignment = left_align
        
        c2.font = regular_font
        c2.border = cell_border
        c2.alignment = right_align if isinstance(kpi_row["Value"], (int, float)) else center_align
        if isinstance(kpi_row["Value"], int):
            c2.number_format = "#,##0"
        curr_row += 1

    # Section 2: Category Breakdown Table
    curr_row += 2
    cat_title = f"Category Analysis (Grouped by '{category_col.replace('_', ' ').title()}')" if category_col else "Category Analysis"
    ws_sum.cell(row=curr_row, column=1, value=cat_title).font = Font(name="Calibri", size=12, bold=True, color="2F5597")
    
    curr_row += 1
    cat_headers = list(category_summary_df.columns)
    for c_idx, h in enumerate(cat_headers, 1):
        cell = ws_sum.cell(row=curr_row, column=c_idx, value=h)
        cell.fill = section_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = header_border
    ws_sum.row_dimensions[curr_row].height = 24

    cat_header_row_idx = curr_row
    curr_row += 1

    for _, cat_row in category_summary_df.iterrows():
        ws_sum.row_dimensions[curr_row].height = 20
        is_grand_total = (cat_row.get("Category") == "Grand Total")
        
        for c_idx, h in enumerate(cat_headers, 1):
            val = cat_row[h]
            cell = ws_sum.cell(row=curr_row, column=c_idx, value=val)
            
            if is_grand_total:
                cell.fill = total_fill
                cell.font = bold_font
                cell.border = total_border
            else:
                cell.font = regular_font
                cell.border = cell_border

            if isinstance(val, (int, float)):
                cell.alignment = right_align
                h_lower = str(h).lower()
                if "total" in h_lower or "price" in h_lower or "amount" in h_lower or "sales" in h_lower or "avg" in h_lower:
                    cell.number_format = "$#,##0.00"
                elif isinstance(val, float):
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "#,##0"
            else:
                cell.alignment = left_align if c_idx == 1 else center_align

        curr_row += 1

    auto_fit_columns(ws_sum)

    # Save final workbook
    wb.save(output_path)


def auto_fit_columns(ws):
    """Auto-fit worksheet column widths based on maximum content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if cell.number_format and "$" in cell.number_format:
                val_str += "    "  # padding for currency signs
            max_len = max(max_len, len(val_str))
        
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def main():
    args = parse_arguments()
    input_path = args.input

    # 1. Validate Input File
    validate_input_filepath(input_path)

    # 2. Determine Output Path
    if args.output:
        output_path = args.output
    else:
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        output_path = os.path.join(os.path.dirname(input_path) or ".", f"{base_name}_report.xlsx")

    print(f"Reading input file: {input_path}")
    raw_df = read_data_file(input_path)

    # 3. Standardize Headers
    df = standardize_columns(raw_df.copy())

    # 4. Auto-detect or Validate Category Column
    category_col = detect_category_column(df, user_category_col=args.category_col)
    print(f"Standardized columns: {list(df.columns)}")
    print(f"Selected category column for summary: '{category_col}'")

    # 5. Clean & Validate Data
    cleaned_df, errors_df, stats = clean_and_validate(df, required_cols_input=args.required_cols)

    # 6. Generate Summaries
    kpi_df, category_summary_df = generate_summary_dataframes(cleaned_df, stats, category_col)

    # 7. Write to Excel & Format
    print(f"Generating Excel report: {output_path}")
    create_formatted_excel(output_path, cleaned_df, errors_df, kpi_df, category_summary_df, category_col)

    # 8. Console Report Summary
    print("\n" + "=" * 55)
    print(" EXCEL REPORT AUTOMATION COMPLETED")
    print("=" * 55)
    print(f"  Input File:          {input_path}")
    print(f"  Output Report:       {output_path}")
    print(f"  Total Raw Rows:      {stats['total_raw_rows']}")
    print(f"  Duplicates Removed:  {stats['duplicate_rows_count']}")
    print(f"  Valid Cleaned Rows:  {stats['cleaned_rows_count']}")
    print(f"  Flagged Error Rows:  {stats['error_rows_count']}")
    print("-" * 55)
    print("  Sheets Created:")
    print("   1. Cleaned Data  (Valid transactions)")
    print("   2. Errors        (Flagged transactions with error reasons)")
    print("   3. Summary       (KPI metrics & Category aggregations)")
    print("=" * 55)
    print(f"\n{stats['cleaned_rows_count']} rows cleaned, {stats['error_rows_count']} rows flagged as errors, output saved to {output_path}\n")


if __name__ == "__main__":
    main()
